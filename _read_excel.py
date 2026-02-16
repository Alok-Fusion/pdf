import openpyxl
import sys

files = [
    "Takeoff_SUNO Nursing and Health Building.xlsx",
    "Takeoff-Health Partners Lakeview -19oct205.xlsx"
]

for f in files:
    print(f"\n{'='*60}")
    print(f"FILE: {f}")
    print('='*60)
    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        print(f"Sheet names: {wb.sheetnames}")
        for sname in wb.sheetnames:
            ws = wb[sname]
            print(f"\n--- Sheet: {sname} ---")
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                if row_count == 0:
                    print(f"Headers: {list(row)}")
                elif row_count <= 15:
                    print(f"  Row {row_count}: {list(row)}")
                row_count += 1
            print(f"  Total rows: {row_count}")
        wb.close()
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
